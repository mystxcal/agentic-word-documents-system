from __future__ import annotations

import importlib.metadata
import hashlib
import subprocess
import zipfile
from collections import Counter
from pathlib import Path

from docx.image.image import Image
from lxml import etree
from openpyxl import load_workbook

from .diagnostics import DiagnosticBag
from .errors import DocumentSystemError
from .rendering import pdf_information, poppler_status
from .sources.markdown_ast import parse_document
from .word.ooxml import content_control_inventory
from .word.package import (
    DOCUMENT_PART,
    DocxPackage,
    XNS,
    core_property_inventory,
    page_furniture_inventory,
    validate_docx,
)


def _integer_attribute(node, namespace: str, name: str) -> int | None:
    if node is None:
        return None
    value = node.get(f"{{{namespace}}}{name}")
    try:
        return int(value) if value is not None else None
    except ValueError:
        return None


def inspect_docx(path: Path) -> dict:
    """Describe a donor or output without assuming a document type or domain."""
    path = Path(path).resolve()
    if not path.is_file():
        raise DocumentSystemError(f"Word file does not exist: {path}")
    validation = validate_docx(path)
    if not validation["valid"]:
        return {"path": str(path), "package_validation": validation}

    package = DocxPackage(path)
    document_root = package.xml(DOCUMENT_PART)
    sections = []
    for index, section in enumerate(document_root.xpath(".//w:sectPr", namespaces=XNS), 1):
        page_size = section.find("w:pgSz", namespaces=XNS)
        margins = section.find("w:pgMar", namespaces=XNS)
        sections.append(
            {
                "section": index,
                "page_width_twips": _integer_attribute(page_size, XNS["w"], "w"),
                "page_height_twips": _integer_attribute(page_size, XNS["w"], "h"),
                "orientation": page_size.get(f"{{{XNS['w']}}}orient") if page_size is not None else None,
                "margins_twips": {
                    name: _integer_attribute(margins, XNS["w"], name)
                    for name in ("top", "right", "bottom", "left", "header", "footer", "gutter")
                },
            }
        )

    controls = {}
    for part_name in [
        DOCUMENT_PART,
        *sorted(name for name in package.parts if name.startswith("word/header") and name.endswith(".xml")),
        *sorted(name for name in package.parts if name.startswith("word/footer") and name.endswith(".xml")),
    ]:
        inventory = content_control_inventory(package.xml(part_name))
        if inventory:
            controls[part_name] = inventory

    style_summary = {"defined": 0, "custom": 0, "ids": []}
    if "word/styles.xml" in package.parts:
        styles_root = package.xml("word/styles.xml")
        styles = styles_root.xpath("./w:style", namespaces=XNS)
        style_summary = {
            "defined": len(styles),
            "custom": sum(1 for style in styles if style.get(f"{{{XNS['w']}}}customStyle") == "1"),
            "ids": [style.get(f"{{{XNS['w']}}}styleId") for style in styles],
        }

    return {
        "path": str(path),
        "size_bytes": path.stat().st_size,
        "package_validation": validation,
        "core_properties": core_property_inventory(package),
        "body": {
            "paragraphs": len(document_root.xpath(".//w:body/w:p", namespaces=XNS)),
            "tables": len(document_root.xpath(".//w:body/w:tbl", namespaces=XNS)),
            "drawings": len(document_root.xpath(".//w:body//w:drawing | .//w:body//w:pict", namespaces=XNS)),
        },
        "sections": sections,
        "page_furniture": page_furniture_inventory(package),
        "content_controls": controls,
        "styles": style_summary,
    }


def _hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _inspect_excel(path: Path) -> dict:
    workbook = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    try:
        sheets = []
        table_suggestions = []
        formula_count = 0
        formula_examples = []
        for worksheet in workbook.worksheets:
            tables = []
            for name, table in worksheet.tables.items():
                reference = table.ref if hasattr(table, "ref") else str(table)
                tables.append({"name": name, "range": reference})
                table_suggestions.append(
                    {
                        "sheet": worksheet.title,
                        "table": name,
                        "component": {
                            "type": "table",
                            "ownership": "source",
                            "source": path.name,
                            "options": {
                                "locator": {"sheet": worksheet.title, "table": name},
                                "view": {"columns": "*", "style_role": "technical"},
                            },
                        },
                    }
                )
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        formula_count += 1
                        if len(formula_examples) < 20:
                            formula_examples.append(
                                {"sheet": worksheet.title, "cell": cell.coordinate, "formula": cell.value}
                            )
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "tables": tables,
                }
            )
    finally:
        workbook.close()
    return {
        "kind": "excel",
        "sheets": sheets,
        "formula_count": formula_count,
        "formula_examples": formula_examples,
        "table_suggestions": table_suggestions,
    }


def _inspect_markdown(path: Path) -> dict:
    diagnostics = DiagnosticBag()
    text = path.read_text(encoding="utf-8-sig")
    parsed = parse_document(text, diagnostics, str(path), strict=False)
    counts = Counter(block.kind for block in parsed.blocks)
    headings = [
        {
            "level": block.value["level"],
            "text": block.value["text"],
            "id": block.block_id,
            "line": block.start_line,
        }
        for block in parsed.blocks
        if block.kind == "heading"
    ]
    return {
        "kind": "markdown_plus",
        "line_count": len(text.splitlines()),
        "block_count": len(parsed.blocks),
        "block_types": dict(sorted(counts.items())),
        "headings": headings,
        "insert_slots": parsed.slot_names,
        "diagnostics": diagnostics.to_list(),
        "component": {
            "type": "document",
            "ownership": "source",
            "source": path.name,
        },
    }


def _inspect_image(path: Path) -> dict:
    try:
        image = Image.from_file(path)
        facts = {
            "pixels": {"width": image.px_width, "height": image.px_height},
            "dpi": {"horizontal": image.horz_dpi, "vertical": image.vert_dpi},
            "content_type": image.content_type,
        }
    except Exception as exc:
        facts = {"inspection_error": str(exc)}
    return {
        "kind": "figure",
        **facts,
        "component": {
            "type": "figure",
            "ownership": "snapshot",
            "source": path.name,
            "alignment": "center",
            "alt_text": None,
        },
    }


def _inspect_visio(path: Path) -> dict:
    pages = []
    try:
        with zipfile.ZipFile(path) as archive:
            root = etree.fromstring(archive.read("visio/pages/pages.xml"))
            for page in root.xpath(".//*[local-name()='Page']"):
                pages.append(
                    {
                        "id": page.get("ID"),
                        "name": page.get("Name"),
                        "universal_name": page.get("NameU"),
                        "background": page.get("Background") == "1",
                    }
                )
    except (OSError, KeyError, zipfile.BadZipFile, etree.XMLSyntaxError) as exc:
        raise DocumentSystemError(f"Could not inspect Visio package {path}: {exc}") from exc
    return {
        "kind": "visio",
        "page_count": len(pages),
        "pages": pages,
        "adapter_status": "inspection_available; Word insertion adapter not yet implemented",
    }


def inspect_source(path: Path) -> dict:
    """Inspect any supported canonical input and expose copyable structural facts."""

    path = Path(path).resolve()
    if not path.is_file():
        raise DocumentSystemError(f"Source file does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix == ".docx":
        detail = {"kind": "word", **inspect_docx(path)}
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        detail = _inspect_excel(path)
    elif suffix == ".pdf":
        pdf = pdf_information(path)
        detail = {
            "kind": "pdf",
            "page_count": pdf["info"].get("pages"),
            "pdf_info": pdf["info"],
            "component_requirements": {
                "type": "pdf_pages",
                "ownership": "source",
                "source": path.name,
                "required_option": "Select explicit one-based source pages before insertion.",
            },
        }
    elif suffix in {".md", ".markdown"}:
        detail = _inspect_markdown(path)
    elif suffix in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff"}:
        detail = _inspect_image(path)
    elif suffix == ".vsdx":
        detail = _inspect_visio(path)
    else:
        raise DocumentSystemError(
            f"No source inspector is registered for {suffix or '(no extension)'}. "
            "Supported sources: DOCX, Markdown, Excel OOXML, PDF, raster images, and VSDX."
        )
    return {
        "schema": "agentic-source-inspection/v1",
        "path": str(path),
        "name": path.name,
        "size_bytes": path.stat().st_size,
        "sha256": _hash(path),
        **detail,
    }


def _word_automation_status() -> dict:
    command = [
        "powershell.exe",
        "-NoLogo",
        "-NoProfile",
        "-Command",
        "$t=[type]::GetTypeFromProgID('Word.Application'); if($null -eq $t){exit 3}; $t.FullName",
    ]
    completed = subprocess.run(command, capture_output=True, text=True, timeout=20, check=False)
    return {
        "available": completed.returncode == 0,
        "detail": (completed.stdout or completed.stderr).strip() or None,
    }


def doctor_status() -> dict:
    from . import __version__

    packages = {}
    for name in ("lxml", "openpyxl", "Pillow", "python-docx", "pydantic"):
        try:
            packages[name] = importlib.metadata.version(name)
        except importlib.metadata.PackageNotFoundError:
            packages[name] = None
    word = _word_automation_status()
    poppler = poppler_status()
    checks = {
        "python_packages": all(packages.values()),
        "word_automation": word["available"],
        "pdf_rendering": poppler["available"],
    }
    return {
        "schema": "agentic-doctor/v2",
        "system_version": __version__,
        "ready": all(checks.values()),
        "checks": checks,
        "packages": packages,
        "word": word,
        "poppler": poppler,
    }
