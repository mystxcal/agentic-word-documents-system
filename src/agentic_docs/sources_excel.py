"""Read an arbitrary Excel table or explicit range into a neutral table dataset.

The reader intentionally knows nothing about BOQs, schedules, registers, or any
other business schema.  Column selection, filtering, sorting, grouping, and row
identity are optional view instructions supplied by the document recipe.
"""

from __future__ import annotations

import hashlib
import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def canonical_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def object_hash(value: Any) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest().upper()


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _matrix(worksheet, cell_range: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return [
        [canonical_value(worksheet.cell(row=row, column=column).value) for column in range(min_col, max_col + 1)]
        for row in range(min_row, max_row + 1)
    ]


def _extract(workbook, locator: dict) -> tuple[list[str], list[list[Any]], dict]:
    table_name = locator.get("table")
    if table_name:
        matches = []
        for worksheet in workbook.worksheets:
            if table_name in worksheet.tables:
                matches.append((worksheet, worksheet.tables[table_name]))
        if len(matches) != 1:
            raise ValueError(f"Expected exactly one Excel Table named {table_name!r}; found {len(matches)}")
        worksheet, table = matches[0]
        values = _matrix(worksheet, table.ref)
        resolved = {"kind": "table", "sheet": worksheet.title, "table": table_name, "range": table.ref}
    else:
        sheet_name = locator.get("sheet")
        cell_range = locator.get("range")
        if not sheet_name or not cell_range:
            raise ValueError("A table source requires locator.table or both locator.sheet and locator.range")
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        values = _matrix(worksheet, cell_range)
        resolved = {"kind": "range", "sheet": sheet_name, "range": cell_range}

    if not values:
        raise ValueError("The selected Excel source is empty")
    has_headers = locator.get("has_headers", True)
    supplied_headers = locator.get("headers")
    if has_headers:
        headers = [str(value).strip() for value in values[0]]
        rows = values[1:]
    else:
        if supplied_headers:
            headers = [str(value).strip() for value in supplied_headers]
        else:
            headers = [f"Column {index + 1}" for index in range(len(values[0]))]
        rows = values
    if any(not header for header in headers):
        raise ValueError("Source headers must not be blank")
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise ValueError(f"Source headers must be unique; duplicates: {', '.join(duplicates)}")
    normalized = [(row + [""] * len(headers))[: len(headers)] for row in rows]
    normalized = [row for row in normalized if not all(value == "" for value in row)]
    return headers, normalized, resolved


def _formula_evidence(workbook_path: Path, resolved: dict) -> dict:
    formulas = load_workbook(workbook_path, data_only=False, read_only=False, keep_links=True)
    cached = load_workbook(workbook_path, data_only=True, read_only=False, keep_links=True)
    try:
        formula_sheet = formulas[resolved["sheet"]]
        cached_sheet = cached[resolved["sheet"]]
        min_col, min_row, max_col, max_row = range_boundaries(resolved["range"])
        count = 0
        missing_cached = []
        examples = []
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                formula_cell = formula_sheet.cell(row=row, column=column)
                if formula_cell.data_type != "f":
                    continue
                count += 1
                if len(examples) < 20:
                    examples.append({"cell": formula_cell.coordinate, "formula": formula_cell.value})
                if cached_sheet.cell(row=row, column=column).value is None:
                    missing_cached.append(formula_cell.coordinate)
        return {
            "formula_count": count,
            "formula_examples": examples,
            "missing_cached_result_count": len(missing_cached),
            "missing_cached_result_cells": missing_cached[:100],
        }
    finally:
        formulas.close()
        cached.close()


def _resolve_columns(headers: list[str], view: dict) -> list[dict]:
    configured = view.get("columns", "*")
    if configured in (None, "*"):
        return [{"source": header, "heading": header} for header in headers]
    if not isinstance(configured, list) or not configured:
        raise ValueError("view.columns must be '*' or a non-empty list")
    result = []
    for item in configured:
        column = {"source": item, "heading": item} if isinstance(item, str) else dict(item)
        source = column.get("source")
        if source not in headers:
            raise ValueError(f"Selected column not found: {source}")
        column.setdefault("heading", source)
        result.append(column)
    return result


def _equal(left: Any, right: Any) -> bool:
    if isinstance(left, str) or isinstance(right, str):
        return str(left).strip().casefold() == str(right).strip().casefold()
    return left == right


def _matches(value: Any, rule: dict) -> bool:
    operation = rule.get("op", "eq")
    expected = rule.get("value")
    if operation == "eq":
        return _equal(value, expected)
    if operation == "ne":
        return not _equal(value, expected)
    if operation == "in":
        return any(_equal(value, item) for item in (expected or []))
    if operation == "not_in":
        return not any(_equal(value, item) for item in (expected or []))
    if operation == "contains":
        return str(expected).casefold() in str(value).casefold()
    if operation == "not_blank":
        return value not in (None, "")
    if operation == "blank":
        return value in (None, "")
    raise ValueError(f"Unsupported filter operation: {operation}")


def _sort_value(value: Any) -> tuple[int, Any]:
    if value in (None, ""):
        return (2, "")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (0, float(value))
    return (1, str(value).casefold())


def export_dataset(
    workbook_path: Path,
    locator: dict,
    view: dict | None = None,
    *,
    component_id: str | None = None,
    formula_policy: str = "cached_values",
) -> dict:
    workbook_path = Path(workbook_path).resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError("The native reader supports OOXML Excel files (.xlsx/.xlsm/.xltx/.xltm)")
    workbook = load_workbook(workbook_path, data_only=True, read_only=False, keep_links=True)
    try:
        headers, rows, resolved = _extract(workbook, locator)
    finally:
        workbook.close()
    formula_evidence = _formula_evidence(workbook_path, resolved)
    if formula_policy == "require_no_formulas" and formula_evidence["formula_count"]:
        raise ValueError(
            f"Selected Excel source contains {formula_evidence['formula_count']} formula cell(s), "
            "but formula_policy is require_no_formulas"
        )
    if formula_policy == "require_cached_results" and formula_evidence["missing_cached_result_count"]:
        cells = ", ".join(formula_evidence["missing_cached_result_cells"][:10])
        raise ValueError(
            f"Selected Excel source has formula cells without cached results: {cells}. "
            "Open, recalculate, and save the workbook in Excel before rebuilding."
        )
    if formula_policy not in {"cached_values", "require_no_formulas", "require_cached_results"}:
        raise ValueError(f"Unsupported Excel formula policy: {formula_policy}")
    index = {header: offset for offset, header in enumerate(headers)}
    view = view or {}
    columns = _resolve_columns(headers, view)

    referenced = [rule.get("column") for rule in view.get("filters", [])]
    referenced += [rule.get("column") for rule in view.get("sort", [])]
    referenced += [view.get("group_by"), view.get("row_id")]
    missing = sorted({name for name in referenced if name and name not in index})
    if missing:
        raise ValueError(f"View references source columns that do not exist: {', '.join(missing)}")

    selected = [
        (source_offset, row)
        for source_offset, row in enumerate(rows, start=1)
        if all(_matches(row[index[rule["column"]]], rule) for rule in view.get("filters", []))
    ]
    for rule in reversed(view.get("sort", [])):
        reverse = rule.get("direction", "asc").lower() == "desc"
        selected.sort(key=lambda item, name=rule["column"]: _sort_value(item[1][index[name]]), reverse=reverse)

    seen_ids: set[str] = set()
    records = []
    for display_index, (source_offset, row) in enumerate(selected, start=1):
        row_id = str(row[index[view["row_id"]]]).strip() if view.get("row_id") else None
        if view.get("row_id") and not row_id:
            raise ValueError(f"Selected source row {source_offset} has a blank row ID in {view['row_id']}")
        if row_id and row_id in seen_ids:
            raise ValueError(f"Duplicate row ID after filtering: {row_id}")
        if row_id:
            seen_ids.add(row_id)
        records.append(
            {
                "id": row_id,
                "group": row[index[view["group_by"]]] if view.get("group_by") else None,
                "values": [row[index[column["source"]]] for column in columns],
                "source_row_offset": source_offset,
                "display_index": display_index,
            }
        )

    dataset = {
        "schema": "agentic-generic-table-dataset/v1",
        "component_id": component_id,
        "source_workbook": str(workbook_path),
        "source_workbook_sha256": file_hash(workbook_path),
        "source_locator": resolved,
        "source_headers": headers,
        "source_nonblank_row_count": len(rows),
        "selected_row_count": len(records),
        "columns": columns,
        "row_id_column": view.get("row_id"),
        "group_by_column": view.get("group_by"),
        "formula_policy": formula_policy,
        "formula_evidence": formula_evidence,
        "records": records,
    }
    dataset["dataset_hash"] = object_hash(
        {"columns": columns, "records": [{key: value for key, value in record.items() if key != "source_row_offset"} for record in records]}
    )
    return dataset
